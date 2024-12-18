import openpyxl
from openpyxl.chart import PieChart, Reference

# Загрузка существующего Excel файла
wb = openpyxl.load_workbook("employee_salaries.xlsx")
ws = wb.active

# Извлечение данных о зарплатах по отделам
departments = {}

# Считывание данных из таблицы (начиная с 2 строки, потому что в первой строке заголовки)
for row in ws.iter_rows(min_row=2, min_col=1, max_col=6, values_only=True):
    if isinstance(row[5], (int, float)):  # Проверка на числовое значение зарплаты
        department = row[2]  # Отдел
        salary = row[5]      # Зарплата
        if department not in departments:
            departments[department] = 0
        departments[department] += salary

# Добавляем отделы с нулевой зарплатой, если они отсутствуют
all_departments = ['Бухгалтерия', 'Отдел кадров', 'Столовая']
for dept in all_departments:
    if dept not in departments:
        departments[dept] = 0

# Подготовка данных для диаграммы
dept_names = list(departments.keys())
dept_salaries = list(departments.values())

# # Проверка корректности данных (выводим их)
# print("Данные по отделам и зарплатам:")
# for dept, salary in zip(dept_names, dept_salaries):
#     print(f"Отдел: {dept}, Зарплата: {salary}")

# Запись данных о зарплатах по отделам в Excel
start_row = ws.max_row + 2  # Начало записи данных для диаграммы после таблицы
ws.cell(row=start_row, column=1, value="Отдел")
ws.cell(row=start_row, column=2, value="Зарплата")

# Запись данных о зарплатах по отделам в Excel
for i, dept in enumerate(dept_names):
    ws.cell(row=start_row + i + 1, column=1, value=dept)
    ws.cell(row=start_row + i + 1, column=2, value=dept_salaries[i])

# Создание круговой диаграммы

print(f"Data range: {start_row + 1} to {start_row + len(dept_names)}")
print(f"Categories range: {start_row + 1} to {start_row + len(dept_names)}")

# Создание диаграммы
data = Reference(ws, min_col=2, min_row=start_row + 1, max_row=start_row + len(dept_names))
categories = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(dept_names))

# Убедимся, что ссылки корректны
# print("Data:", [cell.value for cell in ws.iter_rows(min_col=2, min_row=start_row + 1, max_row=start_row + len(dept_names), values_only=True)])
# print("Categories:", [cell.value for cell in ws.iter_rows(min_col=1, min_row=start_row + 1, max_row=start_row + len(dept_names), values_only=True)])

chart = PieChart()
chart.add_data(data, titles_from_data=True)
chart.set_categories(categories)
chart.title = "Распределение зарплаты по отделам"

# Вставка диаграммы в Excel
ws.add_chart(chart, "E2")  # Диаграмма будет вставлена начиная с ячейки E2

# Сохранение обновленного Excel файла
wb.save("employee_salaries_with_chart.xlsx")
