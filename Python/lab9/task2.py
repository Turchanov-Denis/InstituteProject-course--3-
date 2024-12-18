import openpyxl

# Load the existing Excel workbook
wb = openpyxl.load_workbook("employee_salaries.xlsx")
ws = wb.active

# Read data starting from row 14, columns 2 and 3 (Department and Salary)
data = []
for row in ws.iter_rows(min_row=14, min_col=2, max_col=3, values_only=True):
    data.append(row)

# Print department and salary for each employee
print("Department and salary of employees:")
for row in data:
    print(f"Department: {row[0]}, Salary: {row[1]}")

# Prepare employee data, checking that salary is a number
employee_data = []
for row in ws.iter_rows(min_row=2, min_col=1, max_col=6, values_only=True):
    if isinstance(row[5], (int, float)):  # Check if salary is a numeric value
        employee_data.append({'department': row[2], 'salary': row[5]})

# Find the employee with the maximum and minimum salary
max_salary = max(employee_data, key=lambda x: x['salary'])
min_salary = min(employee_data, key=lambda x: x['salary'])

# Calculate average salary per department
departments = {}
for employee in employee_data:
    dept = employee['department']
    if dept not in departments:
        departments[dept] = {'total_salary': 0, 'count': 0}
    departments[dept]['total_salary'] += employee['salary']
    departments[dept]['count'] += 1

# Print average salary per department
print("\nAverage salary by department:")
for dept, values in departments.items():
    avg_salary = values['total_salary'] / values['count']
    print(f"Average salary in department {dept}: {avg_salary:.2f} rub.")

# Print the employee with the maximum and minimum salary
print(f"\nEmployee with the highest salary: {max_salary['department']} - {max_salary['salary']} rub.")
print(f"Employee with the lowest salary: {min_salary['department']} - {min_salary['salary']} rub.")
